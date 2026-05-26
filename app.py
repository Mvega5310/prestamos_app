from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from sqlalchemy.orm import subqueryload
from functools import wraps
from datetime import date, datetime
import os

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "dev_local_prestamos_2024")

DATABASE_URL = os.environ.get(
    "DATABASE_URL",
    "sqlite:///" + os.path.join(os.path.dirname(os.path.abspath(__file__)), "prestamos.db")
)
if DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)

app.config["SQLALCHEMY_DATABASE_URI"] = DATABASE_URL
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = "login"
login_manager.login_message = "Inicia sesión para continuar."
login_manager.login_message_category = "warning"


# ── Modelos ───────────────────────────────────────────────────────────────────

class Usuario(UserMixin, db.Model):
    __tablename__ = "usuarios"
    id            = db.Column(db.Integer, primary_key=True)
    username      = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(256), nullable=False)
    nombre        = db.Column(db.String(120))
    rol           = db.Column(db.String(20), default="viewer")  # admin | viewer
    activo        = db.Column(db.Boolean, default=True)

    def set_password(self, pw):
        self.password_hash = generate_password_hash(pw, method="pbkdf2:sha256")

    def check_password(self, pw):
        return check_password_hash(self.password_hash, pw)


class Prestamo(db.Model):
    __tablename__ = "prestamos"
    id          = db.Column(db.Integer, primary_key=True)
    nombre      = db.Column(db.String(120), nullable=False)
    fecha       = db.Column(db.Date, nullable=False)
    capital     = db.Column(db.Integer, nullable=False)
    interes_pct = db.Column(db.Float, default=20.0)
    interes     = db.Column(db.Integer, nullable=False)
    total_pagar = db.Column(db.Integer, nullable=False)
    fecha_vence = db.Column(db.Date)
    estado      = db.Column(db.String(20), default="En curso")
    notas       = db.Column(db.Text)
    visible_cobrador = db.Column(db.Boolean, default=True, nullable=False, server_default="1")
    created_at  = db.Column(db.DateTime, default=datetime.utcnow)
    abonos      = db.relationship("Abono", backref="prestamo", lazy=True,
                                  cascade="all, delete-orphan")

    @property
    def total_abonado(self):
        return sum(a.monto for a in self.abonos)

    @property
    def saldo(self):
        return self.total_pagar - self.total_abonado

    @property
    def dias_vence(self):
        if not self.fecha_vence:
            return None
        return (self.fecha_vence - date.today()).days


class Abono(db.Model):
    __tablename__ = "abonos"
    id          = db.Column(db.Integer, primary_key=True)
    prestamo_id = db.Column(db.Integer, db.ForeignKey("prestamos.id"), nullable=False)
    fecha       = db.Column(db.Date, nullable=False)
    monto       = db.Column(db.Integer, nullable=False)
    notas       = db.Column(db.Text)
    created_at  = db.Column(db.DateTime, default=datetime.utcnow)


class Configuracion(db.Model):
    __tablename__ = "configuracion"
    clave = db.Column(db.String(80), primary_key=True)
    valor = db.Column(db.String(256), nullable=False)


# ── Helpers ───────────────────────────────────────────────────────────────────

@login_manager.user_loader
def load_user(uid):
    return db.session.get(Usuario, int(uid))


def admin_required(f):
    @wraps(f)
    @login_required
    def decorated(*args, **kwargs):
        if current_user.rol != "admin":
            flash("Necesitas permisos de administrador.", "danger")
            return redirect(url_for("dashboard"))
        return f(*args, **kwargs)
    return decorated


def _migrate():
    is_sqlite = "sqlite" in DATABASE_URL
    with db.engine.connect() as conn:
        try:
            if is_sqlite:
                conn.execute(db.text(
                    "ALTER TABLE prestamos ADD COLUMN visible_cobrador INTEGER NOT NULL DEFAULT 1"
                ))
            else:
                conn.execute(db.text(
                    "ALTER TABLE prestamos ADD COLUMN visible_cobrador BOOLEAN NOT NULL DEFAULT TRUE"
                ))
            conn.commit()
        except Exception:
            pass

with app.app_context():
    _migrate()


def get_config(clave, default="0"):
    row = db.session.get(Configuracion, clave)
    return row.valor if row else default

def set_config(clave, valor):
    row = db.session.get(Configuracion, clave)
    if row:
        row.valor = str(valor)
    else:
        db.session.add(Configuracion(clave=clave, valor=str(valor)))
    db.session.commit()


def fmt_cop(n):
    try:
        return f"${int(n):,}".replace(",", ".")
    except Exception:
        return n

app.jinja_env.filters["cop"] = fmt_cop
app.jinja_env.globals["today"] = date.today


@app.context_processor
def inject_globals():
    from datetime import timedelta
    if current_user.is_authenticated:
        today = date.today()
        dias = ['lunes','martes','miércoles','jueves','viernes','sábado','domingo']
        meses = ['enero','febrero','marzo','abril','mayo','junio','julio','agosto',
                 'septiembre','octubre','noviembre','diciembre']
        today_human = f"{dias[today.weekday()]} {today.day} de {meses[today.month-1]}"
        return dict(
            activos_count=Prestamo.query.filter_by(estado="En curso").count(),
            today_iso=today.isoformat(),
            default_vence_iso=(today + timedelta(days=30)).isoformat(),
            today_human=today_human,
        )
    return {}


# Crea las tablas al iniciar (funciona con gunicorn y python directo)
with app.app_context():
    db.create_all()


# ── Setup primer uso ──────────────────────────────────────────────────────────

@app.route("/setup", methods=["GET", "POST"])
def setup():
    if Usuario.query.count() > 0:
        return redirect(url_for("login"))
    if request.method == "POST":
        u = Usuario(
            username=request.form["username"].strip(),
            nombre=request.form["nombre"].strip(),
            rol="admin"
        )
        u.set_password(request.form["password"])
        db.session.add(u)
        db.session.commit()
        flash("Administrador creado. Ya puedes iniciar sesión.", "success")
        return redirect(url_for("login"))
    return render_template("setup.html")


# ── Login / Logout ────────────────────────────────────────────────────────────

@app.route("/login", methods=["GET", "POST"])
def login():
    if current_user.is_authenticated:
        return redirect(url_for("dashboard"))
    if request.method == "POST":
        u = Usuario.query.filter_by(
            username=request.form["username"].strip(), activo=True
        ).first()
        if u and u.check_password(request.form["password"]):
            login_user(u, remember=request.form.get("remember") == "on")
            return redirect(request.args.get("next") or url_for("dashboard"))
        flash("Usuario o contraseña incorrectos.", "danger")
    return render_template("login.html")


@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("login"))


# ── Dashboard ─────────────────────────────────────────────────────────────────

@app.route("/")
@login_required
def dashboard():
    q_activos = Prestamo.query.filter_by(estado="En curso")
    if current_user.rol != "admin":
        q_activos = q_activos.filter_by(visible_cobrador=True)
    activos = (q_activos
               .options(subqueryload(Prestamo.abonos))
               .order_by(Prestamo.fecha_vence.asc().nullslast())
               .all())

    pendiente = sum(p.saldo for p in activos)
    hoy = date.today()

    alertas = [p for p in activos if p.dias_vence is not None and p.dias_vence <= 2]

    hoy_data = {
        "prestamos_hoy": len([p for p in activos if p.dias_vence == 0]),
        "recibido_hoy":  db.session.query(db.func.sum(Abono.monto)).filter(Abono.fecha == hoy).scalar() or 0,
        "abonos_hoy":    Abono.query.filter_by(fecha=hoy).count(),
        "por_cobrar":    db.session.query(db.func.sum(Abono.monto)).filter(Abono.fecha == hoy).scalar() or 0,
    }

    stats = {
        "por_cobrar":     pendiente,
        "activos_count":  len(activos),
        "alertas_count":  len(alertas),
        "vencidos_30plus": len([p for p in activos if p.dias_vence is not None and p.dias_vence < -30]),
    }

    top_morosos = []
    mejores_clientes = []

    if current_user.rol == "admin":
        capital_inicial  = int(get_config("capital_inicial", "0"))
        ganancia_neta    = db.session.query(db.func.sum(Prestamo.interes)).scalar() or 0
        capital_colocado = sum(p.total_pagar for p in activos)
        mes_actual       = hoy.replace(day=1)
        recuperado_mes   = (db.session.query(db.func.sum(Abono.monto))
                           .filter(Abono.fecha >= mes_actual).scalar() or 0)
        pagados_mes_list = (Prestamo.query.filter_by(estado="Pagado")
                           .filter(Prestamo.fecha >= mes_actual).all())

        stats.update({
            "recuperado_mes":    recuperado_mes,
            "ganancia_neta":     ganancia_neta,
            "capital_inicial":   capital_inicial,
            "capital_colocado":  capital_colocado,
            "pagados_mes":       len(pagados_mes_list),
            "pagados_mes_monto": sum(p.total_pagar for p in pagados_mes_list),
        })

        top_morosos = sorted(
            [p for p in activos if p.dias_vence is not None and p.dias_vence < 0],
            key=lambda p: p.saldo, reverse=True
        )[:5]

        from collections import defaultdict as _dd2
        pagados_todos = Prestamo.query.filter_by(estado="Pagado").all()
        clientes = _dd2(lambda: {"prestamos_pagados": 0, "total_cobrado": 0})
        for p in pagados_todos:
            clientes[p.nombre]["prestamos_pagados"] += 1
            clientes[p.nombre]["total_cobrado"] += p.total_pagar
        mejores_clientes = sorted(
            [{"nombre": n, **v} for n, v in clientes.items()],
            key=lambda c: c["prestamos_pagados"], reverse=True
        )[:5]

    return render_template("dashboard.html",
        stats=stats,
        hoy=hoy_data,
        alertas=alertas,
        top_morosos=top_morosos,
        mejores_clientes=mejores_clientes)


# ── Lista préstamos ───────────────────────────────────────────────────────────

@app.route("/prestamos")
@login_required
def lista_prestamos():
    buscar   = request.args.get("q", "").strip()
    page     = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 10, type=int)
    if per_page not in (10, 20, 50):
        per_page = 10

    # ── Vista cobrador: agrupada por persona ─────────────────────────────────
    if current_user.rol != "admin":
        from collections import defaultdict
        import math as _math

        q = (Prestamo.query
             .filter_by(estado="En curso", visible_cobrador=True)
             .options(subqueryload(Prestamo.abonos)))
        if buscar:
            q = q.filter(Prestamo.nombre.ilike(f"%{buscar}%"))
        todos = q.order_by(Prestamo.nombre, Prestamo.fecha.desc()).all()

        grupos = defaultdict(list)
        for p in todos:
            grupos[p.nombre].append(p)

        grouped = []
        for nombre in sorted(grupos):
            loans = grupos[nombre]
            tp = sum(l.total_pagar for l in loans)
            ta = sum(l.total_abonado for l in loans)
            vences = [l.fecha_vence for l in loans if l.fecha_vence]
            dias_list = [l.dias_vence for l in loans if l.dias_vence is not None]
            grouped.append({
                "nombre": nombre,
                "count": len(loans),
                "total_pagar": tp,
                "total_abonado": ta,
                "saldo": tp - ta,
                "proxima_vence": min(vences) if vences else None,
                "min_dias": min(dias_list) if dias_list else None,
            })

        total = len(grouped)
        inicio = (page - 1) * per_page
        pages = max(1, _math.ceil(total / per_page))

        return render_template("prestamos.html",
                               es_cobrador=True,
                               grouped=grouped[inicio: inicio + per_page],
                               total=total,
                               page=page,
                               pages=pages,
                               per_page=per_page,
                               buscar=buscar)

    # ── Vista admin: agrupada por persona ────────────────────────────────────
    from collections import defaultdict as _dd
    import math as _math

    filtro = request.args.get("filtro", "activos")
    q = Prestamo.query.options(subqueryload(Prestamo.abonos))
    if buscar:
        q = q.filter(Prestamo.nombre.ilike(f"%{buscar}%"))
    todos = q.order_by(Prestamo.nombre, Prestamo.fecha.desc()).all()

    grupos = _dd(list)
    for p in todos:
        grupos[p.nombre].append(p)

    grouped_admin = []
    for nombre in sorted(grupos):
        loans = grupos[nombre]
        activos = [l for l in loans if l.estado == "En curso"]
        pagados = [l for l in loans if l.estado == "Pagado"]

        if filtro == "activos" and not activos:
            continue
        if filtro == "pagados" and activos:
            continue

        tp = sum(l.total_pagar for l in activos)
        ta = sum(l.total_abonado for l in activos)
        vences   = [l.fecha_vence for l in activos if l.fecha_vence]
        dias_list = [l.dias_vence for l in activos if l.dias_vence is not None]

        todos_visibles = bool(activos) and all(l.visible_cobrador for l in activos)

        grouped_admin.append({
            "nombre":         nombre,
            "count_activos":  len(activos),
            "count_pagados":  len(pagados),
            "total_pagar":    tp,
            "total_abonado":  ta,
            "saldo":          tp - ta,
            "proxima_vence":  min(vences) if vences else None,
            "min_dias":       min(dias_list) if dias_list else None,
            "todos_visibles": todos_visibles,
        })

    total  = len(grouped_admin)
    inicio = (page - 1) * per_page
    pages  = max(1, _math.ceil(total / per_page))

    return render_template("prestamos.html",
                           es_cobrador=False,
                           grouped_admin=grouped_admin[inicio: inicio + per_page],
                           total=total,
                           page=page,
                           pages=pages,
                           filtro=filtro,
                           buscar=buscar,
                           per_page=per_page)


# ── Nuevo préstamo ────────────────────────────────────────────────────────────

@app.route("/prestamos/nuevo", methods=["GET", "POST"])
@admin_required
def nuevo_prestamo():
    if request.method == "POST":
        nombre = request.form.get("nombre", "").strip()
        if not nombre:
            flash("El nombre del prestatario es requerido.", "danger")
            return redirect(url_for("nuevo_prestamo"))
        capital     = int(request.form["capital"])
        interes_pct = float(request.form.get("interes_pct", 20))
        interes     = round(capital * interes_pct / 100)
        fv_str      = request.form.get("fecha_vence")

        p = Prestamo(
            nombre      = nombre,
            fecha       = date.fromisoformat(request.form["fecha"]),
            capital     = capital,
            interes_pct = interes_pct,
            interes     = interes,
            total_pagar = capital + interes,
            fecha_vence = date.fromisoformat(fv_str) if fv_str else None,
            notas       = request.form.get("notas", "").strip() or None,
        )
        db.session.add(p)
        db.session.commit()
        flash(f"Préstamo de {p.nombre} registrado.", "success")
        return redirect(url_for("lista_prestamos"))

    nombre_param = request.args.get("nombre", "").strip()
    nombres_existentes = [r.nombre for r in
        Prestamo.query.with_entities(Prestamo.nombre).distinct().order_by(Prestamo.nombre).all()]
    return render_template("nuevo_prestamo.html", nombre_param=nombre_param,
                           nombres_existentes=nombres_existentes)


# ── Detalle deudor (proxy: redirige según rol) ────────────────────────────────

@app.route("/persona/<nombre>")
@login_required
def detalle_deudor(nombre):
    if current_user.rol == "admin":
        return redirect(url_for("detalle_persona_admin", nombre=nombre))
    return redirect(url_for("detalle_persona_cobrador", nombre=nombre))


# ── Detalle préstamo ──────────────────────────────────────────────────────────

@app.route("/prestamos/<int:pid>")
@login_required
def detalle_prestamo(pid):
    p = (Prestamo.query
         .options(subqueryload(Prestamo.abonos))
         .get_or_404(pid))
    abonos = sorted(p.abonos, key=lambda a: a.fecha, reverse=True)
    return render_template("detalle_prestamo.html",
        p=p, abonos=abonos, hoy=date.today().isoformat())


# ── Detalle persona (cobrador) ────────────────────────────────────────────────

@app.route("/cobrador/persona/<nombre>")
@login_required
def detalle_persona_cobrador(nombre):
    if current_user.rol == "admin":
        return redirect(url_for("lista_prestamos"))

    prestamos = (Prestamo.query
                 .filter_by(nombre=nombre, estado="En curso", visible_cobrador=True)
                 .options(subqueryload(Prestamo.abonos))
                 .order_by(Prestamo.fecha.desc())
                 .all())

    if not prestamos:
        flash("No hay préstamos activos visibles para esta persona.", "warning")
        return redirect(url_for("lista_prestamos"))

    total_pagar   = sum(p.total_pagar for p in prestamos)
    total_abonado = sum(p.total_abonado for p in prestamos)
    saldo_total   = total_pagar - total_abonado

    return render_template("detalle_persona_cobrador.html",
                           nombre=nombre,
                           prestamos=prestamos,
                           total_pagar=total_pagar,
                           total_abonado=total_abonado,
                           saldo_total=saldo_total)


# ── Detalle persona (admin) ───────────────────────────────────────────────────

@app.route("/admin/persona/<nombre>")
@admin_required
def detalle_persona_admin(nombre):
    prestamos = (Prestamo.query
                 .filter_by(nombre=nombre)
                 .options(subqueryload(Prestamo.abonos))
                 .order_by(Prestamo.fecha.desc())
                 .all())

    if not prestamos:
        flash("No hay préstamos para esta persona.", "warning")
        return redirect(url_for("lista_prestamos"))

    activos = [p for p in prestamos if p.estado == "En curso"]
    pagados = [p for p in prestamos if p.estado == "Pagado"]
    tp = sum(p.total_pagar for p in activos)
    ta = sum(p.total_abonado for p in activos)

    return render_template("detalle_persona_admin.html",
                           nombre=nombre,
                           prestamos=prestamos,
                           activos=activos,
                           pagados=pagados,
                           total_pagar=tp,
                           total_abonado=ta,
                           saldo=tp - ta,
                           hoy=date.today().isoformat())


# ── Registrar abono ───────────────────────────────────────────────────────────

@app.route("/prestamos/<int:pid>/abono", methods=["POST"])
@admin_required
def registrar_abono(pid):
    p = Prestamo.query.options(subqueryload(Prestamo.abonos)).get_or_404(pid)
    monto = int(request.form["monto"])

    if monto > p.saldo:
        flash(f"El abono ({fmt_cop(monto)}) supera el saldo ({fmt_cop(p.saldo)}).", "warning")
        return redirect(url_for("detalle_prestamo", pid=pid))

    a = Abono(
        prestamo_id = pid,
        fecha       = date.fromisoformat(request.form["fecha"]),
        monto       = monto,
        notas       = request.form.get("notas", "").strip() or None,
    )
    db.session.add(a)

    if p.saldo - monto == 0:
        p.estado = "Pagado"
        flash(f"Abono registrado. ¡Préstamo de {p.nombre} completamente pagado!", "success")
    else:
        flash(f"Abono de {fmt_cop(monto)} registrado. Saldo: {fmt_cop(p.saldo - monto)}.", "success")

    db.session.commit()
    return redirect(url_for("detalle_prestamo", pid=pid))


# ── Editar abono ──────────────────────────────────────────────────────────────

@app.route("/abonos/<int:aid>/editar", methods=["POST"])
@admin_required
def editar_abono(aid):
    a = Abono.query.get_or_404(aid)
    monto_nuevo = int(request.form["monto"])
    otros_abonos = sum(x.monto for x in a.prestamo.abonos if x.id != a.id)
    if monto_nuevo < 1 or otros_abonos + monto_nuevo > a.prestamo.total_pagar:
        flash("Monto inválido: supera el total a pagar.", "warning")
        return redirect(url_for("detalle_prestamo", pid=a.prestamo_id))
    a.fecha = date.fromisoformat(request.form["fecha"])
    a.monto = monto_nuevo
    a.notas = request.form.get("notas", "").strip() or None
    if a.prestamo.estado == "Pagado" and a.prestamo.saldo != 0:
        a.prestamo.estado = "En curso"
    elif a.prestamo.estado == "En curso" and a.prestamo.saldo == 0:
        a.prestamo.estado = "Pagado"
    db.session.commit()
    flash("Abono actualizado.", "success")
    return redirect(url_for("detalle_prestamo", pid=a.prestamo_id))


@app.route("/abonos/<int:aid>/eliminar", methods=["POST"])
@admin_required
def eliminar_abono(aid):
    a = Abono.query.get_or_404(aid)
    pid = a.prestamo_id
    db.session.delete(a)
    p = Prestamo.query.get(pid)
    if p.estado == "Pagado" and p.saldo != 0:
        p.estado = "En curso"
    db.session.commit()
    flash("Abono eliminado.", "success")
    return redirect(url_for("detalle_prestamo", pid=pid))


# ── Editar préstamo ───────────────────────────────────────────────────────────

@app.route("/prestamos/<int:pid>/editar", methods=["GET", "POST"])
@admin_required
def editar_prestamo(pid):
    p = Prestamo.query.get_or_404(pid)
    if request.method == "POST":
        capital     = int(request.form["capital"])
        interes_pct = float(request.form.get("interes_pct", 20))
        interes     = round(capital * interes_pct / 100)
        p.nombre      = request.form["nombre"].strip()
        p.fecha       = date.fromisoformat(request.form["fecha"])
        fv_str        = request.form.get("fecha_vence")
        p.fecha_vence = date.fromisoformat(fv_str) if fv_str else None
        p.capital     = capital
        p.interes_pct = interes_pct
        p.interes     = interes
        p.total_pagar = capital + interes
        p.estado      = request.form.get("estado", "En curso")
        p.notas       = request.form.get("notas", "").strip() or None
        db.session.commit()
        flash("Préstamo actualizado.", "success")
        return redirect(url_for("detalle_prestamo", pid=pid))
    return render_template("editar_prestamo.html", p=p)


@app.route("/prestamos/<int:pid>/eliminar", methods=["POST"])
@admin_required
def eliminar_prestamo(pid):
    p = Prestamo.query.get_or_404(pid)
    nombre = p.nombre
    db.session.delete(p)
    db.session.commit()
    flash(f"Préstamo de {nombre} eliminado.", "success")
    return redirect(url_for("lista_prestamos"))


# ── Reportes ──────────────────────────────────────────────────────────────────

@app.route("/reportes")
@admin_required
def reportes():
    from sqlalchemy import text
    is_sqlite = "sqlite" in DATABASE_URL
    fmt_mes_p = "strftime('%Y-%m', p.fecha)" if is_sqlite else "to_char(p.fecha, 'YYYY-MM')"
    q_persona = request.args.get("q_persona", "").strip()

    # ── Cuadre por fechas ─────────────────────────────────────────────────────
    desde_str = request.args.get("desde", "")
    hasta_str = request.args.get("hasta", "")
    cuadre_abonos = []
    cuadre_total  = 0
    if desde_str and hasta_str:
        try:
            desde_d = date.fromisoformat(desde_str)
            hasta_d = date.fromisoformat(hasta_str)
            cuadre_abonos = (Abono.query
                .join(Prestamo)
                .filter(Abono.fecha >= desde_d, Abono.fecha <= hasta_d)
                .order_by(Abono.fecha.asc(), Abono.id.asc())
                .all())
            cuadre_total = sum(a.monto for a in cuadre_abonos)
        except ValueError:
            pass

    with db.engine.connect() as conn:
        # ── Por mes ───────────────────────────────────────────────────────────
        por_mes_raw = conn.execute(text(f"""
            SELECT {fmt_mes_p} AS mes,
                   COUNT(DISTINCT p.id) AS cantidad,
                   SUM(p.capital) AS capital,
                   SUM(p.total_pagar) AS total,
                   COALESCE(SUM(a.monto), 0) AS cobrado
            FROM prestamos p
            LEFT JOIN abonos a ON a.prestamo_id = p.id
            GROUP BY mes ORDER BY mes DESC
        """)).mappings().all()

        _meses_es = {1:'ene',2:'feb',3:'mar',4:'abr',5:'may',6:'jun',
                     7:'jul',8:'ago',9:'sep',10:'oct',11:'nov',12:'dic'}
        por_mes = []
        for m in por_mes_raw:
            d = dict(m)
            try:
                parts = d["mes"].split("-")
                d["mes_short"] = _meses_es.get(int(parts[1]), parts[1])
                d["prestamos"] = d["cantidad"]
            except Exception:
                d["mes_short"] = d["mes"]
                d["prestamos"] = d.get("cantidad", 0)
            por_mes.append(d)

        # ── Por prestatario ───────────────────────────────────────────────────
        where_per = ""
        q_per_val = None
        if q_persona:
            where_per = "WHERE p.nombre ILIKE :q_per" if not is_sqlite else "WHERE p.nombre LIKE :q_per"
            q_per_val = f"%{q_persona}%"

        por_prestatario_raw = conn.execute(text(f"""
            SELECT p.nombre,
                   COUNT(DISTINCT p.id) AS veces,
                   SUM(p.capital) AS capital,
                   SUM(p.total_pagar) AS total_pagar,
                   COALESCE(SUM(a.monto_total), 0) AS cobrado,
                   SUM(p.total_pagar) - COALESCE(SUM(a.monto_total), 0) AS pendiente,
                   SUM(CASE WHEN p.estado = 'Pagado' THEN 1 ELSE 0 END) AS prestamos_pagados
            FROM prestamos p
            LEFT JOIN (
                SELECT prestamo_id, SUM(monto) AS monto_total FROM abonos GROUP BY prestamo_id
            ) a ON a.prestamo_id = p.id
            {where_per}
            GROUP BY p.nombre
            ORDER BY pendiente DESC, capital DESC
        """), {"q_per": q_per_val} if q_per_val else {}).mappings().all()
        por_prestatario = [dict(r) for r in por_prestatario_raw]

    total_capital = db.session.query(db.func.sum(Prestamo.capital)).scalar() or 0
    total_interes = db.session.query(db.func.sum(Prestamo.interes)).scalar() or 0
    total_cobrado = db.session.query(db.func.sum(Abono.monto)).scalar() or 0
    total_n       = Prestamo.query.count()

    stats = {
        "prestamos_total": total_n,
        "capital_total":   total_capital,
        "intereses_total": total_interes,
        "cobrado_total":   total_cobrado,
    }

    return render_template("reportes.html",
        stats=stats,
        por_mes=por_mes,
        por_prestatario=por_prestatario,
        q_persona=q_persona,
        desde=desde_str, hasta=hasta_str,
        cuadre_abonos=cuadre_abonos, cuadre_total=cuadre_total)


# ── API autocomplete nombres ──────────────────────────────────────────────────

@app.route("/api/nombres")
@login_required
def api_nombres():
    q = request.args.get("q", "").strip()
    rows = (Prestamo.query
            .with_entities(Prestamo.nombre)
            .filter(Prestamo.nombre.ilike(f"%{q}%"))
            .distinct()
            .order_by(Prestamo.nombre)
            .limit(10)
            .all())
    return jsonify([r.nombre for r in rows])


# ── API info persona (préstamos activos) ──────────────────────────────────────

@app.route("/api/persona_info")
@login_required
def api_persona_info():
    nombre = request.args.get("nombre", "").strip()
    if not nombre:
        return jsonify(None)
    prestamos = (Prestamo.query
                 .filter_by(nombre=nombre, estado="En curso")
                 .options(subqueryload(Prestamo.abonos))
                 .order_by(Prestamo.fecha.desc())
                 .all())
    if not prestamos:
        return jsonify(None)
    tp = sum(p.total_pagar for p in prestamos)
    ta = sum(p.total_abonado for p in prestamos)
    return jsonify({
        "count": len(prestamos),
        "total_pagar": tp,
        "total_abonado": ta,
        "saldo": tp - ta,
        "interes_pct": prestamos[0].interes_pct,
    })


# ── Exportar Excel ───────────────────────────────────────────────────────────

@app.route("/exportar")
@login_required
def exportar_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    import io

    wb = openpyxl.Workbook()

    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill("solid", fgColor="1a2340")
    center       = Alignment(horizontal="center")

    def estilizar(ws, headers):
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.font, c.fill, c.alignment = header_font, header_fill, center
        ws.row_dimensions[1].height = 18

    def autoajustar(ws):
        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    def cop(n):
        try: return f"${int(n):,}".replace(",", ".")
        except: return n

    prestamos = (Prestamo.query
                 .filter_by(estado="En curso")
                 .options(subqueryload(Prestamo.abonos))
                 .order_by(Prestamo.fecha.desc(), Prestamo.id.desc()).all())

    # ── Hoja 1: Deudores activos ─────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Deudores activos"
    headers1 = ["#", "Nombre", "Fecha", "Capital", "Interés %", "Total a pagar",
                "Abonado", "Saldo", "Fecha vence"]
    estilizar(ws1, headers1)
    for p in prestamos:
        ws1.append([
            p.id, p.nombre,
            p.fecha.strftime("%d/%m/%Y"),
            cop(p.capital), p.interes_pct,
            cop(p.total_pagar), cop(p.total_abonado),
            cop(p.saldo),
            p.fecha_vence.strftime("%d/%m/%Y") if p.fecha_vence else "",
        ])
    autoajustar(ws1)

    # ── Hoja 2: Abonos ───────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Abonos")
    headers2 = ["# Préstamo", "Nombre", "Fecha abono", "Monto", "Notas"]
    estilizar(ws2, headers2)
    abonos = (Abono.query
              .join(Prestamo)
              .order_by(Abono.fecha.desc()).all())
    for a in abonos:
        ws2.append([
            a.prestamo_id, a.prestamo.nombre,
            a.fecha.strftime("%d/%m/%Y"),
            cop(a.monto), a.notas or ""
        ])
    autoajustar(ws2)

    # ── Hoja 3: Por prestatario ──────────────────────────────────────────────
    ws3 = wb.create_sheet("Por prestatario")
    headers3 = ["Nombre", "Veces", "Capital total", "Total a pagar", "Cobrado", "Pendiente"]
    estilizar(ws3, headers3)
    from sqlalchemy import text
    is_sqlite = "sqlite" in DATABASE_URL
    with db.engine.connect() as conn:
        rows = conn.execute(text("""
            SELECT p.nombre,
                   COUNT(*) AS veces,
                   SUM(p.capital) AS capital_total,
                   SUM(p.total_pagar) AS total_pagar,
                   COALESCE(SUM(a.abonado),0) AS abonado,
                   SUM(p.total_pagar) - COALESCE(SUM(a.abonado),0) AS pendiente
            FROM prestamos p
            LEFT JOIN (
                SELECT prestamo_id, SUM(monto) AS abonado
                FROM abonos GROUP BY prestamo_id
            ) a ON a.prestamo_id = p.id
            GROUP BY p.nombre ORDER BY pendiente DESC
        """)).mappings().all()
    for r in rows:
        ws3.append([r.nombre, r.veces, cop(r.capital_total),
                    cop(r.total_pagar), cop(r.abonado), cop(r.pendiente)])
    autoajustar(ws3)

    # ── Hoja 4: Por mes ──────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Por mes")
    headers4 = ["Mes", "Préstamos", "Capital", "Total"]
    estilizar(ws4, headers4)
    fmt_mes = "strftime('%Y-%m', fecha)" if is_sqlite else "to_char(fecha, 'YYYY-MM')"
    with db.engine.connect() as conn:
        rows = conn.execute(text(f"""
            SELECT {fmt_mes} AS mes, COUNT(*) AS cantidad,
                   SUM(capital) AS capital, SUM(total_pagar) AS total
            FROM prestamos GROUP BY mes ORDER BY mes DESC
        """)).mappings().all()
    for r in rows:
        ws4.append([r.mes, r.cantidad, cop(r.capital), cop(r.total)])
    autoajustar(ws4)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nombre_archivo = f"Kuenta_{date.today().strftime('%Y%m%d')}.xlsx"
    return send_file(buf, as_attachment=True,
                     download_name=nombre_archivo,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ── Gestión de usuarios (solo admin) ─────────────────────────────────────────

@app.route("/usuarios")
@admin_required
def lista_usuarios():
    usuarios = Usuario.query.order_by(Usuario.username).all()
    return render_template("usuarios.html", usuarios=usuarios)


@app.route("/usuarios/nuevo", methods=["GET", "POST"])
@admin_required
def nuevo_usuario():
    if request.method == "POST":
        username = request.form["username"].strip()
        if Usuario.query.filter_by(username=username).first():
            flash("Ese nombre de usuario ya existe.", "warning")
            return redirect(url_for("nuevo_usuario"))
        u = Usuario(
            username = username,
            nombre   = request.form["nombre"].strip(),
            rol      = request.form.get("rol", "viewer"),
        )
        u.set_password(request.form["password"])
        db.session.add(u)
        db.session.commit()
        flash(f"Usuario '{u.username}' creado.", "success")
        return redirect(url_for("lista_usuarios"))
    return render_template("nuevo_usuario.html")


@app.route("/usuarios/<int:uid>/toggle", methods=["POST"])
@admin_required
def toggle_usuario(uid):
    u = db.session.get(Usuario, uid)
    if u and u.id != current_user.id:
        u.activo = not u.activo
        db.session.commit()
        flash(f"Usuario '{u.username}' {'activado' if u.activo else 'desactivado'}.", "success")
    return redirect(url_for("lista_usuarios"))


@app.route("/usuarios/<int:uid>/reset", methods=["POST"])
@admin_required
def reset_password(uid):
    u = db.session.get(Usuario, uid)
    nueva = request.form.get("password", "").strip()
    if u and nueva:
        u.set_password(nueva)
        db.session.commit()
        flash(f"Contraseña de '{u.username}' actualizada.", "success")
    return redirect(url_for("lista_usuarios"))


# ── Visibilidad cobrador ──────────────────────────────────────────────────────

@app.route("/prestamos/<int:pid>/visibilidad", methods=["POST"])
@admin_required
def toggle_visibilidad(pid):
    p = Prestamo.query.get_or_404(pid)
    p.visible_cobrador = not p.visible_cobrador
    db.session.commit()
    estado = "visible" if p.visible_cobrador else "oculto"
    flash(f"Préstamo de {p.nombre} ahora es {estado} para el cobrador.", "success")
    return redirect(url_for("detalle_prestamo", pid=pid))


@app.route("/admin/persona/<nombre>/visibilidad", methods=["POST"])
@admin_required
def toggle_visibilidad_persona(nombre):
    activos = Prestamo.query.filter_by(nombre=nombre, estado="En curso").all()
    if not activos:
        return redirect(url_for("lista_prestamos"))
    # Si todos están visibles → ocultar todos. Si alguno está oculto → mostrar todos.
    todos_visibles = all(p.visible_cobrador for p in activos)
    nuevo_estado = not todos_visibles
    for p in activos:
        p.visible_cobrador = nuevo_estado
    db.session.commit()
    estado_txt = "visible" if nuevo_estado else "oculto"
    flash(f"{nombre} ahora es {estado_txt} para el cobrador.", "success")
    return redirect(url_for("lista_prestamos"))


# ── Ajustes (solo admin) ─────────────────────────────────────────────────────

@app.route("/ajustes", methods=["GET", "POST"])
@admin_required
def ajustes():
    if request.method == "POST":
        raw = request.form.get("capital_inicial", "0").replace(".", "").replace(",", "").strip()
        set_config("capital_inicial", int(raw) if raw.isdigit() else 0)
        flash("Ajustes guardados.", "success")
        return redirect(url_for("ajustes"))
    capital_inicial = int(get_config("capital_inicial", "0"))
    return render_template("ajustes.html", capital_inicial=capital_inicial)


# ── Perfil ────────────────────────────────────────────────────────────────────

@app.route("/perfil", methods=["GET", "POST"])
@login_required
def perfil():
    if request.method == "POST":
        accion = request.form.get("accion")

        if accion == "nombre":
            current_user.nombre = request.form["nombre"].strip()
            db.session.commit()
            flash("Nombre actualizado.", "success")

        elif accion == "password":
            actual = request.form["password_actual"]
            nueva  = request.form["password_nueva"]
            confirmar = request.form["password_confirmar"]
            if not current_user.check_password(actual):
                flash("La contraseña actual es incorrecta.", "danger")
            elif nueva != confirmar:
                flash("Las contraseñas nuevas no coinciden.", "warning")
            elif len(nueva) < 6:
                flash("La contraseña debe tener mínimo 6 caracteres.", "warning")
            else:
                current_user.set_password(nueva)
                db.session.commit()
                flash("Contraseña actualizada correctamente.", "success")

        return redirect(url_for("perfil"))

    return render_template("perfil.html")


# ── Init ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    debug = os.environ.get("FLASK_DEBUG", "0") == "1"
    port = int(os.environ.get("PORT", "5050"))
    app.run(debug=debug, port=port)
